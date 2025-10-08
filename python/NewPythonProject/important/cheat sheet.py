# an expression is a piece of code that returns a value
# a variable is something we use to store a value
# string is a piece of writing --> "hello"
# integer is a number without decimal and float is a number with decimal
# Boolean is a value that can only be True or False
from ecommers.shipping import calc_shipping
from important.convertors import lbs_to_kg

# input:
name = input('what is your name? ')
# print:
print(name + "  Hello world " * 10)

# int converts string to integer
# float converts string into float
# bool converts string into boolean

# if we want a string that goes into multiple lines we must use ''' ''' or """ """
print('''Hi 
I am
Alive
''')

# with [] we choose the index
course = "My life"
# in this case indexes are: 0 1 2 3 4 5
# also negative index start from the end---> [-1] = e
print(course[0])
# in print(course[0:3]) we can print indexes from 0 to 2
# if we want to start from the beginning or end of an index we can keep that index empty : [:5]

# everything inside a formatted string is considered a string until you tell it otherwise
# this helps you write a string without dealing with different characters
my_name = "siavash"
msg = f'{my_name} is beautiful ):"><'
print(msg)

# len function can count the number of characters in an object:
print(len("doli"))

# the 'in' operator allows for boolean value
print("A" in course)

# in python: - minus + plus * times / divide ** to the power
# methods are commands we can use:
print(course.upper())

# by using math module by typing:
import math

# we can use its methods:
print(math.ceil(2.9))

# if statements use boolean values:
is_cold = True
is_hot = False

if is_hot:
    print("Drink")
# elif stands for else if
elif is_cold:
    print("Warm")
else:
    print("Run")

# there are other operators as well such as:
# == which allows you to check for equality in an if statement
if my_name == "siavash":
    print("Yes, it’s Siavash")

# "and" allows you to choose multiple conditions:
if is_hot and is_cold:
    print("Impossible!")

# "not" allows you to reverse a value
if is_hot and not is_cold:
    print("Hot but not cold")

# > and < and >= and <= which allow comparisons
temp = 15
if temp > 12:
    print("Temperature is greater than 12")

# you can use while statement to make loops
# break allows for terminating a loop
i = 1
while i <= 5:
    print(i)
    i = i + 1
    break

# a for loop is used when you want to repeat something a fixed number of times or go through a sequence (like a list, string, or range of numbers).
for item in "spain":
    print(item)

# you can use [] to make a list
example_list = [1, 2, 3, 4, "bowling"]

# you can use range function to set from one number to another
# the first number is where it starts, the second one where it ends, and the third is step
for item in range(5, 20, 3):
    print(item)

# in math a matrix is just a way of arranging numbers in rows and columns, like a table.
list1 = [1, 2, 3,
         4, 5, 6,
         7, 8, 9]

# we can model list using two-dimensional lists which is essentially lists inside lists
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

# we can access a matrix like this
print(matrix[0][1])

# you can use methods to add, remove or modify a list
numbers = [1, 2, 3]
numbers.append(4)  # adds an item to the end of the list
numbers.insert(0, 10)  # allows you to add an item wherever in the list
numbers.remove(2)  # removes a specific item
numbers.pop()  # removes the last item on a list
print(numbers.index(3))  # allows to search for index of items
numbers.sort()  # sorts the list from least to most
numbers2 = numbers.copy()  # now numbers2 is a copy of numbers
print(numbers2)

# you can use .split to split a value into list:
msg = "good morning all"
msg_list = msg.split(" ")  # split for each space
print(msg_list)

# tuple is a list that cannot be modified
tuple1 = (1, 2, 3)
# tuple1[1] = 10  # ❌ gives an error because tuples are immutable

# each item in a list can be given its own value:
coordinate = [1, 2, 3]
x = coordinate[0]
y = coordinate[1]
z = coordinate[2]
print(x, y, z)

# or we can use unpacking feature:
x, y, z = coordinate
print(x, y, z)

# a dictionary is a built-in data type that stores data as key–value pairs
customer = {
    "name": "jj",
    "email": "@jj.com",
    "number": 1234
}
print(customer["name"])


# function is a container for a few lines of code that do a specific task
# it breaks our code into smaller chunks
def greet_user():
    print("hello there!")
    print("welcome!")


# call the function
print("start")
greet_user()
print("finish")


# we can set parameter for our functions to change on demand
def send_metal(metal):
    print(f"sending {metal}")


print("welcome to mine")
send_metal("silver")  # now silver will replace metal variable


# we can also send a value and get a value back but we must use a return statement
# you can only send one return value
def emoji_converter(message):
    emojis = {
        ":)": "🙂",
        ":(": "😥"
    }
    output = ""
    message_list = message.split(" ")
    for word in message_list:
        output += emojis.get(word, word) + " "
    return output


message = input(">")
print(emoji_converter(message))

# you can use try-except blocks to handle different types of error
try:
    age = int(input(">"))
    print(age)
except ValueError:
    print("wrong value")
# until now we have used different classes such as string
# but we can actually define our own classes with their own methods
class Point:
    #now we define the methods
    def move(self):
        print("move")

    def draw(self):
        print("draw")
#now we need to create a new object witch can be defined as our new class
point1 = Point()
point1.move()
#these objects can also have an attribute witch is a value that belongs to that object
point1.x = 10
#we can add define the attribute during the creation of the object
#for this purpose we add a method called the constructor
class bake:
    def __init__(self,flour,water): #this is the method that gets called when we make a new object
        self.flour = flour
        self.water = water
    def heat(self):
        print("heating")
strawberry = bake(12, 1)
#now we may have multiple classes that use the same methods so we can use inheritance to stop our self's from repeating lines of code
class Mammal:
    def walk(self):
        print("walking")

class Dog(Mammal):  # Dog inherits from Mammal
    pass

class Cat(Mammal):  # Cat also inherits from Mammal
    pass


dino = Dog()
kitty = Cat()

dino.walk()
kitty.walk()
#we can break our code down to multiple files in order to make our codes smaller and more reusable
#now we import the functions from convertors.py
import convertors
heavy = 60
new_heavy =convertors.lbs_to_kg(heavy)
print(new_heavy)
#or we can import a specific module
from convertors import lbs_to_kg
print(lbs_to_kg(heavy))
# we can make a python package that holds our module
#and now we call that package and the module file
import ecommers.shipping
ecommers.shipping.calc_shipping()
#or we can use the "from" statement
from ecommers.shipping import calc_shipping
calc_shipping()
#or if we want to use the shipping functions
from ecommers import shipping
shipping.calc_shipping()
#python comes with a library of module that dose many things and we can import even more with pip
#when we want to work with exel in python we first import openpyexel and give it a name to make our code shorter
import openpyxl as xl
from openpyxl.chart import  BarChart, Reference #we use this later to make charts
#now we load our work sheet and give it an object
wb = xl.load_workbook("transactions.xlsx")
#in our workbook we have only one sheet and that is sheet one
sheet =wb['Sheet1']
#now to accesses a cell we need to give it coordinate
cell =sheet["a1"]
#we cam also use cell method
cell = sheet.cell(1,1)
#here are some of the useful methods
print(cell.value)#print cell value
print(sheet.max_row)#print the max row in a sheet
#now we make a for loop for each row and modify all rows in colum 3
for row in range(2, sheet.max_row +1):
    cells=sheet.cell(row, 3)
    corrected_price = cells.value * 0.9
#now we add the new price in colum 4
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value= corrected_price
wb.save('transactions2.xlsx')
#now we use the Reference class to select a range of values
values = Reference(sheet, min_row=2,
                 max_row=sheet.max_row,
                 min_col=4,
                 max_col=4)
chart = BarChart() #now we make a BarChart
chart.add_data(values) #we add the data to it
sheet.add_chart(chart, 'E2') # now we add the chart to the sheet and specifie where it should go

wb.save("transactions2.xlsx")