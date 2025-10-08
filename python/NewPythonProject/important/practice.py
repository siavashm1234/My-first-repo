import openpyxl as xl
from openpyxl.chart import  BarChart, Reference #we use this later to make charts
#now we load our work sheet and give it an object
wb = xl.load_workbook("transactions.xlsx")
#in our workbook we have only one sheet and that is sheet one
sheet =wb['Sheet1']
#now to accesses a cell we need to give it coordinate
cell =sheet["a1"]
#we cam also use cell method
cell = sheet.cell(1,1)
#here are some of the useful methods
print(cell.value)#print cell value
print(sheet.max_row)#print the max row in a sheet
#now we make a for loop for each row and modify all rows in colum 3
for row in range(2, sheet.max_row +1):
    cells=sheet.cell(row, 3)
    corrected_price = cells.value * 0.9
#now we add the new price in colum 4
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value= corrected_price
wb.save('transactions2.xlsx')
#now we use the Reference class to select a range of values
values = Reference(sheet, min_row=2,
                 max_row=sheet.max_row,
                 min_col=4,
                 max_col=4)
chart = BarChart() #now we make a BarChart
chart.add_data(values) #we add the data to it
sheet.add_chart(chart, 'E2') # now we add the chart to the sheet and specifie where it should go
wb.save("transactions2.xlsx")